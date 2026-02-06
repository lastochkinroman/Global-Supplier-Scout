import os
from datetime import datetime
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from config import Config
from database import product_db, Product


class ExcelReportGenerator:
    def __init__(self):
        self.reports_dir = Config.TEMP_DIR
        os.makedirs(self.reports_dir, exist_ok=True)

        self.header_fill = PatternFill(
            start_color="CCCCCC",
            end_color="CCCCCC",
            fill_type="solid"
        )
        self.header_font = Font(bold=True, size=10)
        self.center_alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )
        self.left_alignment = Alignment(
            horizontal="left",
            vertical="center",
            wrap_text=True
        )
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def generate_supplier_analysis_report(self, products_data: List[Dict[str, Any]]) -> str:
        wb = Workbook()
        ws = wb.active
        ws.title = "Анализ поставщиков"

        self._add_report_header(ws)
        self._add_data_headers(ws)
        self._populate_report_data(ws, products_data)
        self._auto_resize_columns(ws)
        self._add_summary_sheet(wb, products_data)

        return self._save_report(wb)

    def _add_report_header(self, ws):
        ws.merge_cells('A1:Z1')
        title_cell = ws['A1']
        title_cell.value = (
            f"Маркетинговое исследование: Отчет анализа поставщиков\n"
            f"Сгенерирован: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        )
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

    def _add_data_headers(self, ws):
        headers = [
            "№", "Код товара", "Название товара", "Полное название товара",
            "Категория", "Единица", "Единица в документе", "Базовая цена (USD)",
            "Поставщик", "Страна поставщика", "Рейтинг поставщика",
            "Цена USD", "Цена RUB", "Доставка %", "Доставка RUB",
            "Хранение %", "Хранение RUB", "Название дополнительных затрат",
            "Дополнительные затраты %", "Дополнительные затраты RUB",
            "Конечная цена RUB", "Конечная цена USD", "Время доставки",
            "МОК (USD)", "Место склада", "Статус поставщика",
            "Сайт поставщика", "ИНН поставщика", "Вес товара (кг)",
            "Размеры (см)", "Год", "Квартал"
        ]

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.thin_border

    def _populate_report_data(self, ws, products_data: List[Dict[str, Any]]):
        row_idx = 4
        current_year = datetime.now().year
        current_quarter = (datetime.now().month - 1) // 3 + 1

        for product_data in products_data:
            product: Product = product_data["product"]
            suppliers = product_data["suppliers"][:Config.MAX_SUPPLIERS_PER_PRODUCT]

            for supplier_idx, supplier in enumerate(suppliers, 1):
                product_code = product_db.generate_product_code(product, supplier)

                row_data = [
                    supplier_idx,
                    product_code,
                    product.name,
                    product.full_name,
                    product.category,
                    product.unit,
                    product.doc_unit,
                    product.base_price_usd,
                    supplier["name"],
                    supplier["country"],
                    supplier["rating"],
                    supplier["price_usd"],
                    supplier["price_rub"],
                    supplier["delivery_cost_percent"],
                    supplier["delivery_cost_rub"],
                    supplier["storage_cost_percent"],
                    supplier["storage_cost_rub"],
                    supplier["additional_costs_name"],
                    supplier["additional_costs_percent"],
                    supplier["additional_costs_rub"],
                    supplier["final_price_rub"],
                    supplier["final_price_usd"],
                    supplier["lead_time"],
                    supplier["moq"],
                    supplier["warehouse_location"],
                    supplier["status"],
                    supplier["url"],
                    supplier["tax_id"],
                    product.weight_kg,
                    product.dimensions_cm,
                    current_year,
                    current_quarter
                ]

                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = self.left_alignment
                    cell.border = self.thin_border

                    if col_idx in [8, 11, 12, 14, 15, 16, 17, 19, 20, 21, 22, 24]:
                        if value is not None:
                            cell.number_format = '#,##0.00'

                    if col_idx == 10:
                        if value is not None:
                            cell.number_format = '0.0'

                row_idx += 1

            row_idx += 1

    def _auto_resize_columns(self, ws):
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _add_summary_sheet(self, wb, products_data: List[Dict[str, Any]]):
        ws = wb.create_sheet(title="Сводка")

        ws.merge_cells('A1:E1')
        title_cell = ws['A1']
        title_cell.value = "Сводка анализа поставщиков"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center")

        headers = ["Товар", "Лучший поставщик", "Лучшая цена (USD)", "Время доставки", "Рейтинг"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill

        row_idx = 4
        for product_data in products_data:
            product = product_data["product"]
            suppliers = product_data["suppliers"]

            if suppliers:
                best_supplier = min(suppliers, key=lambda x: x["final_price_usd"])

                row_data = [
                    product.name,
                    best_supplier["name"],
                    best_supplier["final_price_usd"],
                    best_supplier["lead_time"],
                    best_supplier["rating"]
                ]

                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)

                    if col_idx == 3:
                        cell.number_format = '#,##0.00'
                    elif col_idx == 5:
                        cell.number_format = '0.0'

                row_idx += 1

        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _save_report(self, wb: Workbook) -> str:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"анализ_поставщиков_{timestamp}.xlsx"
        filepath = os.path.join(self.reports_dir, filename)
        wb.save(filepath)
        return filepath


report_generator = ExcelReportGenerator()